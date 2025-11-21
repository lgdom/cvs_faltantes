import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from openpyxl.drawing.image import Image  # <--- NUEVO IMPORT

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Faltantes App", page_icon="📝", layout="wide")

# --- ARCHIVOS EN GITHUB ---
FILE_CLIENTES = 'clientes.csv'
FILE_PRODUCTOS = 'productos.csv'
FILE_PLANTILLA = 'plantilla.xlsx'
FILE_IMAGEN = 'logo.png'

# --- ESTADO DE LA APP ---
if 'pedidos' not in st.session_state:
    st.session_state.pedidos = []
if 'carrito' not in st.session_state:
    st.session_state.carrito = []

# --- CARGA DE DATOS ---
@st.cache_data
def cargar_datos():
    errores = []
    df_cli = pd.DataFrame()
    df_prod = pd.DataFrame()
    
    # 1. CARGAR CLIENTES
    try:
        # CORRECCIÓN: Intentar UTF-8 PRIMERO para respetar acentos
        try:
            df_cli = pd.read_csv(FILE_CLIENTES, encoding='utf-8')
        except UnicodeDecodeError:
            # Si falla, usar latin-1 (Windows antiguo)
            df_cli = pd.read_csv(FILE_CLIENTES, encoding='latin-1')
            
        df_cli.columns = df_cli.columns.str.strip().str.upper()
        
        col_clave = next((c for c in df_cli.columns if 'CLAVE' in c or 'CODIGO' in c), df_cli.columns[0])
        col_nombre = next((c for c in df_cli.columns if ('CLIENTE' in c or 'NOMBRE' in c) and c != col_clave), df_cli.columns[1])
        
        df_cli = df_cli[[col_clave, col_nombre]].copy()
        df_cli.columns = ['CODIGO', 'NOMBRE']
        df_cli['DISPLAY'] = df_cli['CODIGO'].astype(str) + " - " + df_cli['NOMBRE'].astype(str)
        
    except Exception as e:
        errores.append(f"⚠️ Error leyendo Clientes: {e}")
        
        # --- CORRECCIÓN AQUÍ ---
        # 1. Primero buscamos la columna del CÓDIGO (buscando "CLAVE" o "CODIGO")
        col_clave = next((c for c in df_cli.columns if 'CLAVE' in c or 'CODIGO' in c), df_cli.columns[0])
        
        # 2. Luego buscamos la columna del NOMBRE, pero EXCLUYENDO la que ya detectamos como clave
        col_nombre = next((c for c in df_cli.columns if ('CLIENTE' in c or 'NOMBRE' in c) and c != col_clave), df_cli.columns[1])
        # -----------------------
        
        df_cli = df_cli[[col_clave, col_nombre]].copy()
        df_cli.columns = ['CODIGO', 'NOMBRE']
        
        # Convertimos a texto para evitar errores de suma y mostramos "CÓDIGO - NOMBRE"
        df_cli['DISPLAY'] = df_cli['CODIGO'].astype(str) + " - " + df_cli['NOMBRE'].astype(str)
        
    except Exception as e:
        errores.append(f"⚠️ Error leyendo Clientes: {e}")

    # 2. CARGAR PRODUCTOS
    try:
        # --- CORRECCIÓN DE ACENTOS AQUÍ ---
        try:
            # Intento 1: UTF-8 (Estándar moderno, ideal para acentos)
            df_prod = pd.read_csv(FILE_PRODUCTOS, encoding='utf-8')
        except:
            # Intento 2: Latin-1 (Estándar de Excel antiguo/Windows)
            df_prod = pd.read_csv(FILE_PRODUCTOS, encoding='latin-1')
        # ----------------------------------
            
        df_prod.columns = df_prod.columns.str.strip().str.upper()
        
        # Detección de columnas
        col_clave = next(c for c in df_prod.columns if 'CLAVE' in c or 'CODIGO' in c)
        col_desc = next(c for c in df_prod.columns if 'NOMBRE' in c or 'DESCRIPCION' in c)
        # Buscamos sustancia (si no existe, no falla, devuelve None)
        col_sust = next((c for c in df_prod.columns if 'SUSTANCIA' in c), None)
        
        cols_to_keep = [col_clave, col_desc]
        if col_sust: cols_to_keep.append(col_sust)
            
        df_prod = df_prod[cols_to_keep].copy()
        
        # Renombrar estándar
        nombres_std = ['CODIGO', 'DESCRIPCION']
        if col_sust: nombres_std.append('SUSTANCIA')
        df_prod.columns = nombres_std
        
        if 'SUSTANCIA' not in df_prod.columns:
            df_prod['SUSTANCIA'] = '---'
        else:
            df_prod['SUSTANCIA'] = df_prod['SUSTANCIA'].fillna('---')

        # Índice de Búsqueda
        df_prod['SEARCH_INDEX'] = (
            df_prod['CODIGO'].astype(str) + " | " + 
            df_prod['DESCRIPCION'].astype(str) + " | " + 
            df_prod['SUSTANCIA'].astype(str)
        ).str.upper()
        
    except Exception as e:
        errores.append(f"⚠️ Error leyendo Productos: {e}")

    return df_cli, df_prod, errores

df_clientes, df_productos, logs = cargar_datos()

# --- INTERFAZ VISUAL ---
st.title("📝 Registro de Faltantes")

if logs:
    for l in logs: st.error(l)
    st.stop()

with st.sidebar:
    st.info(f"Catálogo cargado: {len(df_productos)} productos")
    if st.button("🗑️ Reiniciar Pedido Actual"):
        st.session_state.carrito = []
        st.rerun()

tab1, tab2 = st.tabs(["1. Registrar Faltantes", "2. Descargar Excel"])

# === PESTAÑA 1: REGISTRO ===
with tab1:
    # --- FUNCIONES CALLBACK (Lógica de limpieza) ---
    def agregar_producto():
        # Recuperamos valores
        cliente = st.session_state.cliente_box
        prod_str = st.session_state.prod_box # Este es el valor del Selectbox único
        cant = st.session_state.qty_box
        
        if cliente and prod_str:
            # Buscar datos
            row = df_productos[df_productos['SEARCH_INDEX'] == prod_str].iloc[0]
            item = {
                "CODIGO": row['CODIGO'],
                "DESCRIPCION": row['DESCRIPCION'],
                "SOLICITADA": cant,
                "SURTIDO": 0,
                "O.C.": "N/A"
            }
            st.session_state.carrito.append(item)
            
            # LIMPIEZA:
            st.session_state.qty_box = 1      
            st.session_state.prod_box = None  # <--- ESTO REINICIA EL BUSCADOR AL FINAL
        else:
            st.warning("⚠️ Selecciona Cliente y Producto primero")

    def finalizar_pedido(fecha):
        if st.session_state.cliente_box:
            cod_cli, nom_cli = st.session_state.cliente_box.split(" - ", 1)
            
            # Guardamos lo que esté en el editor de datos en ese momento
            # Nota: st.session_state.editor_data contiene los cambios
            # Reconstruimos el dataframe final desde el carrito actualizado
            
            pedido_nuevo = {
                "cli_cod": cod_cli,
                "cli_nom": nom_cli,
                "fecha": fecha,
                "items": pd.DataFrame(st.session_state.carrito) # Guardamos estado actual
            }
            st.session_state.pedidos.append(pedido_nuevo)
            
            # LIMPIEZA TOTAL
            st.session_state.carrito = []
            st.session_state.cliente_box = None # Limpiar cliente
            st.session_state.search_box = ""
        else:
            st.error("Falta seleccionar cliente")

    # --- INTERFAZ ---
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("Datos del Pedido")
        
        # A. CLIENTE (Con key)
        st.selectbox(
            "Cliente:", 
            options=df_clientes['DISPLAY'], 
            index=None, 
            placeholder="Escribe para buscar cliente...",
            key="cliente_box"
        )
        fecha_input = st.date_input("Fecha:", datetime.today())
        
        st.divider()
        st.subheader("Agregar Producto")
        
        # B. BUSCADOR UNIFICADO (Tipo Google)
        # Cargamos TODAS las opciones. Streamlit filtra solo al escribir.
        st.selectbox(
            "Buscar Producto (Nombre, Clave o Sustancia):", 
            options=df_productos['SEARCH_INDEX'], # Pasamos la lista completa
            index=None, 
            placeholder="Escribe para buscar...",
            key="prod_box" # Esta key es la que limpiamos en el callback
        )
        
        # C. CANTIDAD
        st.number_input("Cantidad:", min_value=1, value=1, key="qty_box")
        
        # BOTÓN AGREGAR
        st.button("➕ Agregar a la Lista", on_click=agregar_producto, use_container_width=True)
        
    with col2:
        st.subheader("🛒 Lista Preliminar")
        
        if st.session_state.carrito:
            df_cart = pd.DataFrame(st.session_state.carrito)
            
            # TABLA EDITABLE
            df_edited = st.data_editor(
                df_cart,
                column_config={
                    "SOLICITADA": st.column_config.NumberColumn("Solicitada", width="small"),
                    "SURTIDO": st.column_config.NumberColumn("Surtido", width="small"),
                    "O.C.": st.column_config.TextColumn("O.C.", width="small")
                },
                use_container_width=True,
                num_rows="dynamic",
                key="editor_data"
            )
            
            # Sincronizar cambios manuales de la tabla con el carrito en memoria
            # Esto asegura que si editas un número en la tabla, se guarde bien.
            if not df_edited.equals(df_cart):
                st.session_state.carrito = df_edited.to_dict('records')
            
            # BOTÓN GUARDAR (Usa on_click y args para pasar la fecha)
            st.button(
                "💾 GUARDAR ESTE PEDIDO (Siguiente Hoja)", 
                type="primary", 
                use_container_width=True,
                on_click=finalizar_pedido,
                args=(fecha_input,)
            )

        else:
            st.info("👈 Usa el panel izquierdo para buscar productos.")

# === PESTAÑA 2: DESCARGA ===
with tab2:
    st.subheader(f"📦 Pedidos listos para procesar: {len(st.session_state.pedidos)}")
    for i, p in enumerate(st.session_state.pedidos):
        with st.expander(f"Hoja {i+1}: {p['cli_nom']} ({len(p['items'])} prods)"):
            st.dataframe(p['items'])
            if st.button("Eliminar Hoja", key=f"del_{i}"):
                st.session_state.pedidos.pop(i)
                st.rerun()

    st.divider()
    nombre_archivo = st.text_input("Nombre del archivo:", value="Reporte_Faltantes.xlsx")
    
    if st.button("🚀 GENERAR EXCEL FINAL", disabled=(len(st.session_state.pedidos)==0)):
        try:
            wb = openpyxl.load_workbook(FILE_PLANTILLA)
            hoja_base = wb.active
            hoja_base.title = "Base"
            
            conteo_hojas = {}
            for pedido in st.session_state.pedidos:
                cod = pedido['cli_cod']
                conteo_hojas[cod] = conteo_hojas.get(cod, 0) + 1
                nombre_hoja = cod if conteo_hojas[cod] == 1 else f"{cod}-{conteo_hojas[cod]}"
                
                ws = wb.copy_worksheet(hoja_base)
                ws.title = nombre_hoja
                
                ws['B2'] = "SUC. TIJ"
                ws['B3'] = "LUIS FELIPE GARCÍA DOMÍNGUEZ"
                ws['B4'] = pedido['cli_nom']
                try:
                    ws['B6'] = int(cod)
                except:
                    ws['B6'] = cod
                ws['D6'] = pedido['fecha'].strftime('%d/%m/%Y')
                
                fila_inicial = 10
                datos = pedido['items'][['CODIGO', 'DESCRIPCION', 'SOLICITADA', 'SURTIDO', 'O.C.']].values.tolist()
                for idx, fila in enumerate(datos):
                    ws.cell(row=fila_inicial+idx, column=1, value=fila[0])
                    ws.cell(row=fila_inicial+idx, column=2, value=fila[1])
                    ws.cell(row=fila_inicial+idx, column=3, value=fila[2])
                    ws.cell(row=fila_inicial+idx, column=4, value=fila[3])
                    ws.cell(row=fila_inicial+idx, column=5, value=fila[4])
                    
                    # --- RE-INSERTAR IMAGEN ---
                try:
                    img = Image(FILE_IMAGEN)
                    img.width = 200  
                    img.height = 80
                    img.anchor = 'D1' 
                    ws.add_image(img)
                except Exception as e:
                    print(f"No se pudo cargar la imagen en esta hoja: {e}")
            
            del wb['Base']
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                label="⬇️ DESCARGAR ARCHIVO COMPLETO",
                data=buffer,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error generando Excel: {e}")
